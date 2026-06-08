"""Excel export of a filtered Tender queryset using openpyxl.

Produces a polished report:
  * Report title + applied filters + generated date/time + user/company.
  * Bold headers, frozen header row, auto column widths, autofilter.
  * Clickable official source links, readable dates.
Files are saved under exports/excel/ and also returned as bytes for HTTP download.
"""
from __future__ import annotations

import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("tenders")

HEADERS = [
    "Sıra No",
    "İhale Kayıt No",
    "İhale Başlığı",
    "İdare Adı",
    "İl",
    "İlçe",
    "İhale Türü",
    "Kategori",
    "İhale Usulü",
    "İlan Tarihi",
    "İhale Tarihi",
    "Son Teklif Tarihi",
    "İşin Yapılacağı Yer",
    "Kısa Açıklama",
    "Anahtar Kelime Eşleşmesi",
    "Durum",
    "Kaynak",
    "EKAP / Resmi Link",
    "Sisteme Eklenme Tarihi",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_ZEBRA_FILL = PatternFill("solid", fgColor="F2F5F8")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
_META_FONT = Font(italic=True, size=10, color="404040")
_LINK_FONT = Font(color="0563C1", underline="single")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fmt_date(d) -> str:
    return d.strftime("%d.%m.%Y") if d else ""


def build_workbook(tenders, *, filters: Optional[dict] = None, user=None) -> Workbook:
    filters = filters or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "İhaleler"

    ncols = len(HEADERS)
    last_col = get_column_letter(ncols)

    # --- Title ---
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "Türkiye İhale Takip Raporu"
    c.font = _TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # --- Meta lines ---
    meta_lines = []
    company = ""
    if user is not None:
        profile = getattr(user, "profile", None)
        company = getattr(profile, "company_name", "") or user.get_username()
    if company:
        meta_lines.append(f"Firma / Kullanıcı: {company}")
    prov = filters.get("province") or "Tüm Türkiye"
    meta_lines.append(f"İl: {prov}")
    if filters.get("district"):
        meta_lines.append(f"İlçe: {filters['district']}")
    if filters.get("category_label"):
        meta_lines.append(f"Kategori: {filters['category_label']}")
    if filters.get("date_from") or filters.get("date_to"):
        meta_lines.append(
            f"Tarih Aralığı: {filters.get('date_from') or '...'} - "
            f"{filters.get('date_to') or '...'}"
        )
    meta_lines.append(
        f"Oluşturulma: {timezone.localtime():%d.%m.%Y %H:%M}"
    )
    meta_lines.append(f"Toplam Kayıt: {len(tenders)}")

    row = 2
    for line in meta_lines:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        cell = ws[f"A{row}"]
        cell.value = line
        cell.font = _META_FONT
        row += 1

    header_row = row + 1

    # --- Header row ---
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[header_row].height = 30

    # --- Data rows ---
    type_map = dict(__import__("tenders.constants", fromlist=["x"]).TENDER_TYPE_CHOICES)
    cat_map = dict(__import__("tenders.constants", fromlist=["x"]).CATEGORY_CHOICES)
    proc_map = dict(__import__("tenders.constants", fromlist=["x"]).PROCEDURE_CHOICES)
    status_map = dict(__import__("tenders.constants", fromlist=["x"]).STATUS_CHOICES)

    r = header_row + 1
    for idx, t in enumerate(tenders, start=1):
        values = [
            idx,
            t.tender_no,
            t.title,
            t.authority_name,
            t.province,
            t.district,
            type_map.get(t.tender_type, t.tender_type),
            cat_map.get(t.category, t.category),
            proc_map.get(t.tender_procedure, t.tender_procedure),
            _fmt_date(t.announcement_date),
            _fmt_date(t.tender_date),
            _fmt_date(t.deadline_date),
            t.work_location,
            t.short_description,
            t.keyword_matches,
            status_map.get(t.status, t.status),
            t.get_source_display() if hasattr(t, "get_source_display") else t.source,
            t.official_url,
            _fmt_date(getattr(t, "created_at", None).date() if getattr(t, "created_at", None) else None),
        ]
        ws.row_dimensions[r].height = 22
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = _BORDER
            # Center text vertically. Wrap only for long text fields (Title & Authority name)
            cell.alignment = Alignment(vertical="center", wrap_text=col in (3, 4))
            
            # Apply zebra striping to even rows
            if idx % 2 == 0:
                cell.fill = _ZEBRA_FILL
                
            if col == ncols - 1 and val:  # official link column
                cell.hyperlink = val
                cell.font = _LINK_FONT
        r += 1

    # --- Freeze header row ---
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # --- AutoFilter on header + data ---
    ws.auto_filter.ref = f"A{header_row}:{last_col}{max(r - 1, header_row)}"

    # --- Auto column widths ---
    widths = {
        1: 8, 2: 16, 3: 45, 4: 32, 5: 14, 6: 16, 7: 14, 8: 22, 9: 22,
        10: 13, 11: 13, 12: 15, 13: 24, 14: 50, 15: 24, 16: 14, 17: 16,
        18: 30, 19: 18,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    return wb


def export_to_bytes(tenders, *, filters=None, user=None) -> bytes:
    wb = build_workbook(list(tenders), filters=filters, user=user)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_to_file(tenders, *, filters=None, user=None, filename=None) -> Path:
    export_dir = Path(getattr(settings, "EXPORT_DIR"))
    export_dir.mkdir(parents=True, exist_ok=True)
    if not filename:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prov = (filters or {}).get("province") or "Tum_Turkiye"
        safe = "".join(ch for ch in prov if ch.isalnum() or ch in "_-") or "rapor"
        filename = f"ihale_raporu_{safe}_{stamp}.xlsx"
    path = export_dir / filename
    wb = build_workbook(list(tenders), filters=filters, user=user)
    wb.save(path)
    logger.info("Excel exported: %s", path)
    return path
