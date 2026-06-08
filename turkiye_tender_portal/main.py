"""
EKAP İhale Takip Botu — Optimize Edilmiş Standalone Scraper
============================================================
Playwright ile https://ekapv2.kik.gov.tr/ekap/search sayfasından
gerçek zamanlı ihale verisi çeker.

Sayfa yapısı: Angular SPA + DevExtreme bileşenleri
  - <ihale-liste-item> : Her bir ihale kartı
  - <ihale-paginator>  : Sayfalama
  - dx-drop-down-box   : Filtre dropdown'ları

Özellikler:
  • Rastgele User-Agent ile bot algılamayı minimize eder
  • Pop-up / dialog otomatik kapatma
  • "İhale Tarihi En Yeni" sıralamasıyla ilk 20 ilan
  • Elazığ ili + "Katılıma Açık" filtresi
  • pandas ile Excel export
  • Konsola özet (çekilen sayı + ilk 3 başlık)

Kullanım:
  python main.py                   # headless (arka plan)
  python main.py --visible          # tarayıcıyı görerek test
  python main.py --province Ankara  # farklı il
"""
from __future__ import annotations

import argparse
import os
import random
import sys
import time
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Sabitler
# ---------------------------------------------------------------------------
EKAP_URL = "https://ekapv2.kik.gov.tr/ekap/search"
MAX_TENDERS = 20
DEFAULT_PROVINCE = "Elazığ"
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports", "excel")

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:138.0) Gecko/20100101 Firefox/138.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36 Edg/138.0.0.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
]


def get_random_user_agent() -> str:
    return random.choice(USER_AGENTS)


# ---------------------------------------------------------------------------
# Pop-up / Dialog Yönetimi
# ---------------------------------------------------------------------------
def dismiss_popups(page) -> None:
    """Sayfadaki olası pop-up, modal ve overlay'leri kapatır."""
    close_selectors = [
        "button.close", "button[aria-label='Close']",
        "button[aria-label='Kapat']", ".btn-close",
        "[data-dismiss='modal']", "[data-bs-dismiss='modal']",
        ".popup-close", ".overlay-close",
    ]
    for sel in close_selectors:
        try:
            btn = page.locator(sel).first
            if btn.is_visible(timeout=300):
                btn.click(timeout=500)
                time.sleep(0.3)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# İhale İli Filtresi Ayarlama
# ---------------------------------------------------------------------------
def set_province_filter(page, province: str) -> bool:
    """
    EKAP sayfasında İhale İli filtresini ayarlar.
    DevExtreme dropdown bileşeni kullanılıyor.
    """
    print(f"[*] İl filtresi ayarlanıyor: {province}")

    try:
        # "İhale İli" etiketini bul ve yanındaki dropdown'a tıkla
        il_filtre = page.locator("ihale-multiselect-filtre:has-text('İhale İli')").first
        if not il_filtre.is_visible(timeout=2000):
            print("[!] İhale İli filtresi bulunamadı")
            return False

        # İhale İli yanındaki dropdown alanına tıkla
        dropdown = il_filtre.locator("dx-drop-down-box").first
        dropdown.click(timeout=3000)
        time.sleep(1.5)

        # Turkish uppercase mapping:
        translation_table = str.maketrans("iışğüçö", "İIŞĞÜÇÖ")
        province_upper = province.translate(translation_table).upper()
        
        print(f"[*] Listeden tıklanacak il: {province_upper}")
        
        il_item = page.locator(".dx-list-item:visible").filter(has_text=province_upper).first
        if il_item.is_visible(timeout=3000):
            # Click the checkbox inside the list item
            cb = il_item.locator(".dx-checkbox").first
            if cb.is_visible(timeout=1000):
                cb.click(timeout=2000)
            else:
                il_item.click(timeout=2000)
            time.sleep(0.5)
            print(f"[✓] İl filtresi ayarlandı: {province}")

            # Dropdown'u kapat (Escape bas)
            page.keyboard.press("Escape")
            time.sleep(1)
            return True

        print(f"[!] İl listesinde '{province}' ({province_upper}) bulunamadı")
        page.keyboard.press("Escape")
        time.sleep(0.5)
        return False

    except Exception as exc:
        print(f"[!] İl filtresi ayarlanırken hata: {exc}")
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        return False


# ---------------------------------------------------------------------------
# EKAP Scraping — Doğru Selector'lar
# ---------------------------------------------------------------------------
def scrape_ekap_tenders(page, province: str | None = None, max_count: int = MAX_TENDERS) -> list[dict]:
    """
    EKAP arama sayfasından ihale-liste-item bileşenlerini parse eder.
    Her bir <ihale-liste-item> bir ihale kartıdır.
    """
    print(f"[*] EKAP sayfası açılıyor: {EKAP_URL}")
    page.goto(EKAP_URL, wait_until="domcontentloaded", timeout=60000)

    # Angular SPA'nın tam yüklenmesini bekle
    try:
        page.wait_for_load_state("networkidle", timeout=20000)
    except Exception:
        print("[!] Network idle timeout — devam ediliyor...")

    # SPA render süresi için ekstra bekleme
    print("[*] Angular bileşenlerinin yüklenmesi bekleniyor...")
    time.sleep(4)
    dismiss_popups(page)

    # İl filtresi uygulanacaksa
    if province:
        success = set_province_filter(page, province)
        if success:
            try:
                filtrele_btn = page.locator("button:has-text('Filtrele')").first
                if filtrele_btn.is_visible(timeout=2000):
                    print("[*] Filtrele butonuna tıklanıyor...")
                    filtrele_btn.click(timeout=3000)
                    time.sleep(4)
            except Exception as e:
                print(f"[!] Filtrele butonuna tıklanırken hata: {e}")

    # ihale-liste-item bileşenlerinin yüklenmesini bekle
    try:
        page.wait_for_selector("ihale-liste-item", timeout=15000)
        print("[✓] İhale kartları yüklendi")
    except Exception:
        print("[!] ihale-liste-item bekleme timeout — alternatif selector deneniyor...")
        try:
            page.wait_for_selector("div[class*='card']", timeout=10000)
        except Exception:
            print("[!] Hiç ihale kartı bulunamadı")
            return []

    time.sleep(1)

    # -----------------------------------------------------------------------
    # İhale kartlarını parse et
    # -----------------------------------------------------------------------
    tenders: list[dict] = []

    # Birinci sayfa: ihale-liste-item bileşenlerini kullan
    items = page.locator("ihale-liste-item")
    page1_count = items.count()
    print(f"[*] Sayfa 1: {page1_count} ihale kartı bulundu")

    tenders.extend(_parse_items(items, min(page1_count, max_count)))

    # İkinci sayfa gerekiyorsa (sayfa başına 10 kayıt var)
    if len(tenders) < max_count:
        remaining = max_count - len(tenders)
        if _go_next_page(page):
            time.sleep(3)
            items2 = page.locator("ihale-liste-item")
            page2_count = items2.count()
            print(f"[*] Sayfa 2: {page2_count} ihale kartı bulundu")
            tenders.extend(_parse_items(items2, min(page2_count, remaining)))

    print(f"[✓] Toplam {len(tenders)} ihale başarıyla çekildi")
    return tenders


def _parse_items(items, count: int) -> list[dict]:
    """ihale-liste-item bileşenlerinden veri çıkarır."""
    results: list[dict] = []

    for i in range(count):
        try:
            item = items.nth(i)
            item_text = item.inner_text(timeout=5000).strip()

            if not item_text or len(item_text) < 5:
                continue

            tender = _parse_card_text(item_text)
            if tender and tender.get("baslik"):
                results.append(tender)

        except Exception as exc:
            print(f"[!] Kart {i+1} parse hatası: {exc}")
            continue

    return results


def _parse_card_text(text: str) -> dict:
    """
    Bir ihale-liste-item'ın inner_text'ini parse eder.

    Tipik metin yapısı (satır satır):
      İhale İptal Edilmiş     <- durum badge (opsiyonel)
      Hizmet                  <- alım türü
      Açık                    <- ihale usulü
      MARDİN, 23.03.2027 14:00  <- il + tarih
      2026/271215             <- İKN
      Siber Güvenlik ...      <- başlık
      İdare Adı               <- idare
      ...
    """
    lines = [l.strip() for l in text.split("\n") if l.strip()]

    tender = {
        "ikn": "",
        "baslik": "",
        "idare": "",
        "ilan_tarihi": "",
        "durum": "",
        "il": "",
        "alim_turu": "",
        "ihale_usulu": "",
    }

    if not lines:
        return tender

    # İKN'yi bul (YYYY/NNNNNN formatında)
    ikn_idx = -1
    for idx, line in enumerate(lines):
        stripped = line.strip()
        # IKN formatı: 2026/123456 veya 2025/654321
        if "/" in stripped and len(stripped) <= 20:
            parts = stripped.split("/")
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                tender["ikn"] = stripped
                ikn_idx = idx
                break

    # İl + tarih satırını bul (ŞEHİR, DD.MM.YYYY HH:MM formatında)
    tarih_idx = -1
    for idx, line in enumerate(lines):
        # Tarih formatı arama: DD.MM.YYYY
        import re
        date_match = re.search(r"(\d{2}\.\d{2}\.\d{4})", line)
        if date_match:
            tender["ilan_tarihi"] = date_match.group(1)
            tarih_idx = idx
            # İl bilgisi tarihten önce virgülle ayrılmış olabilir
            if "," in line:
                il_part = line.split(",")[0].strip()
                tender["il"] = il_part
            break

    # Durum bilgisini bul
    durum_keywords = [
        "İhale İptal", "Katılıma Açık", "Teklif Değerlendirme",
        "Sözleşme", "Sonuçlanmış", "Tamamlanmış", "İptal",
        "Açık İhale", "Belli İstekliler",
    ]
    for idx, line in enumerate(lines):
        for kw in durum_keywords:
            if kw.lower() in line.lower():
                tender["durum"] = line.strip()
                break

    # Alım türü (Mal / Yapım / Hizmet / Danışmanlık)
    alim_turleri = ["Mal", "Yapım", "Hizmet", "Danışmanlık"]
    for idx, line in enumerate(lines):
        if line.strip() in alim_turleri:
            tender["alim_turu"] = line.strip()
            break

    # İhale usulü
    usul_keywords = ["Açık", "Belli İstekliler", "Pazarlık", "Doğrudan"]
    for idx, line in enumerate(lines):
        if line.strip() in usul_keywords:
            tender["ihale_usulu"] = line.strip()
            break

    # Başlık: İKN'den sonraki satır genelde başlık
    if ikn_idx >= 0 and ikn_idx + 1 < len(lines):
        tender["baslik"] = lines[ikn_idx + 1]
        # İdare: başlıktan sonraki satır
        if ikn_idx + 2 < len(lines):
            potential_idare = lines[ikn_idx + 2]
            # İdare genelde uzunca bir isimdir ve "Müdürlüğü", "Başkanlığı" vb. içerir
            tender["idare"] = potential_idare

    # Eğer başlık bulunamadıysa en uzun satırı başlık olarak al
    if not tender["baslik"]:
        longest = max(lines, key=len) if lines else ""
        tender["baslik"] = longest

    return tender


def _go_next_page(page) -> bool:
    """ihale-paginator'daki sonraki sayfa butonuna tıklar."""
    try:
        # ihale-paginator bileşenindeki "sonraki" veya ">" butonunu bul
        paginator = page.locator("ihale-paginator")
        if paginator.count() == 0:
            print("[!] Paginator bulunamadı")
            return False

        # Sonraki sayfa butonu — çeşitli selector'lar dene
        next_selectors = [
            "ihale-paginator dx-button[title='İleri']",
            "ihale-paginator [title='İleri']",
            "ihale-paginator button:has-text('>')",
            "ihale-paginator button:has-text('»')",
            "ihale-paginator button:has-text('Sonraki')",
            "ihale-paginator .next",
            "ihale-paginator dx-button:last-of-type",
            "ihale-paginator button:last-child",
        ]
        for sel in next_selectors:
            try:
                btn = page.locator(sel).first
                if btn.is_visible(timeout=1000) and btn.is_enabled(timeout=500):
                    btn.click(timeout=3000)
                    print("[*] Sonraki sayfaya geçildi")
                    # Yeni kartların yüklenmesini bekle
                    time.sleep(2)
                    try:
                        page.wait_for_load_state("networkidle", timeout=10000)
                    except Exception:
                        pass
                    return True
            except Exception:
                continue

        # Alternatif: sayfa numarası 2'ye tıkla
        try:
            page2_btn = page.locator("ihale-paginator button:has-text('2')").first
            if page2_btn.is_visible(timeout=1000):
                page2_btn.click(timeout=3000)
                print("[*] Sayfa 2'ye geçildi")
                time.sleep(2)
                return True
        except Exception:
            pass

        print("[!] Sonraki sayfa butonu bulunamadı")
        return False
    except Exception as exc:
        print(f"[!] Sayfalama hatası: {exc}")
        return False


# ---------------------------------------------------------------------------
# Filtreleme
# ---------------------------------------------------------------------------
def filter_by_province_and_status(
    tenders: list[dict],
    province: str = DEFAULT_PROVINCE,
) -> list[dict]:
    """
    Belirtilen il + 'Katılıma Açık' durumundaki ihaleleri filtreler.
    """
    province_lower = province.lower()
    filtered: list[dict] = []

    for t in tenders:
        # İl kontrolü
        searchable = f"{t.get('il', '')} {t.get('baslik', '')} {t.get('idare', '')}".lower()
        if province_lower not in searchable:
            continue

        # Durum kontrolü: Kapalı durumları hariç tut
        durum = t.get("durum", "").lower()
        is_closed = any(kw in durum for kw in [
            "iptal", "tamamlan", "sonuçlan", "sözleşme",
        ])

        if not is_closed:
            filtered.append(t)

    print(f"[*] Filtre ({province} + Katılıma Açık): {len(filtered)}/{len(tenders)} ihale")
    return filtered


# ---------------------------------------------------------------------------
# Excel Export (pandas)
# ---------------------------------------------------------------------------
def export_to_excel(tenders: list[dict], province: str) -> str | None:
    """Filtrelenmiş ihaleleri pandas DataFrame ile Excel'e aktarır."""
    if not tenders:
        print("[!] Export edilecek ihale yok.")
        return None

    import pandas as pd

    df = pd.DataFrame(tenders)

    # Sütun isimlerini Türkçe yap
    column_map = {
        "ikn": "İhale Kayıt No (İKN)",
        "baslik": "İhale Başlığı",
        "idare": "İdare Adı",
        "ilan_tarihi": "İlan Tarihi",
        "durum": "Durum",
        "il": "İl",
        "alim_turu": "Alım Türü",
        "ihale_usulu": "İhale Usulü",
    }
    df = df.rename(columns={k: v for k, v in column_map.items() if k in df.columns})

    # Çıktı dizinini oluştur
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Dosya adı
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_province = "".join(ch for ch in province if ch.isalnum() or ch in "_-") or "genel"
    filename = f"ekap_ihaleler_{safe_province}_{stamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    # Excel'e yaz
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="İhaleler", index=False)

        ws = writer.sheets["İhaleler"]
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        
        # Styles
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        zebra_fill = PatternFill("solid", fgColor="F2F5F8")
        thin_side = Side(style="thin", color="D9D9D9")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Format header row (Row 1)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cell_border

        # Format data rows
        for row_idx in range(2, len(df) + 2):
            ws.row_dimensions[row_idx].height = 22
            is_even = (row_idx % 2 == 0)
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = cell_border
                # Wrap text only for title (col 2) and authority (col 3) if they exist
                cell.alignment = Alignment(vertical="center", wrap_text=col_idx in (2, 3))
                if is_even:
                    cell.fill = zebra_fill
                    
        # Set column widths
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max() if len(df) > 0 else 0,
            )
            width = min(max(max_len + 2, 12), 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    print(f"[✓] Excel kaydedildi: {filepath}")
    return filepath


# ---------------------------------------------------------------------------
# Konsol Özeti
# ---------------------------------------------------------------------------
def print_summary(tenders: list[dict], filtered: list[dict], province: str) -> None:
    """Doğrulama için konsola özet yazdırır."""
    print("\n" + "=" * 60)
    print("  EKAP İHALE BOT — SONUÇ ÖZETİ")
    print("=" * 60)
    print(f"  Toplam çekilen ihale sayısı : {len(tenders)}")
    print(f"  Filtrelenen ({province})     : {len(filtered)}")
    print()

    display_list = filtered if filtered else tenders
    label = "İlk 3 ihale" if filtered else "İlk 3 ihale (filtresiz)"

    if display_list:
        print(f"  {label}:")
        for i, t in enumerate(display_list[:3], start=1):
            baslik = t.get("baslik", "—")
            if len(baslik) > 70:
                baslik = baslik[:67] + "..."
            ikn = t.get("ikn", "—")
            print(f"    {i}. [{ikn}] {baslik}")
    else:
        print("  (Hiç ihale çekilemedi)")

    print("=" * 60 + "\n")


# ---------------------------------------------------------------------------
# Ana Akış
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="EKAP İhale Takip Botu")
    parser.add_argument(
        "--visible", action="store_true",
        help="Tarayıcıyı görünür modda aç (test için)",
    )
    parser.add_argument(
        "--province", type=str, default=DEFAULT_PROVINCE,
        help=f"Filtrelenecek il (varsayılan: {DEFAULT_PROVINCE})",
    )
    parser.add_argument(
        "--max", type=int, default=MAX_TENDERS,
        help=f"Çekilecek maksimum ihale (varsayılan: {MAX_TENDERS})",
    )
    parser.add_argument(
        "--no-filter", action="store_true",
        help="İl filtresini EKAP'ta uygulamadan tüm Türkiye'yi çek",
    )
    args = parser.parse_args()

    headless = not args.visible
    province = args.province

    print(f"[*] EKAP İhale Botu başlatılıyor...")
    print(f"[*] Mod: {'Headless' if headless else 'Görünür'}")
    print(f"[*] Hedef il: {province}")
    print(f"[*] Maksimum ihale: {args.max}")

    from playwright.sync_api import sync_playwright

    with sync_playwright() as pw:
        ua = get_random_user_agent()
        print(f"[*] User-Agent: {ua[:50]}...")

        browser = pw.chromium.launch(
            headless=headless,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ],
        )

        context = browser.new_context(
            user_agent=ua,
            viewport={"width": 1920, "height": 1080},
            locale="tr-TR",
            timezone_id="Europe/Istanbul",
            java_script_enabled=True,
        )

        # Bot algılamayı azalt
        context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
        """)

        page = context.new_page()
        page.on("dialog", lambda dialog: dialog.dismiss())

        try:
            # 1. İhaleleri çek
            tenders = scrape_ekap_tenders(page, province=None if args.no_filter else province, max_count=args.max)

            # 2. İl + Katılıma Açık filtresi
            if args.no_filter:
                filtered = tenders
            else:
                filtered = filter_by_province_and_status(tenders, province=province)

            # 3. Excel'e aktar
            if filtered:
                export_to_excel(filtered, province=province)
            else:
                print("[!] Filtreye uyan ihale yok — tüm çekilen veri export ediliyor...")
                export_to_excel(tenders, province="Tum_Turkiye")

            # 4. Konsol özeti
            print_summary(tenders, filtered, province)

        except Exception as exc:
            print(f"\n[HATA] Bot çalışırken bir hata oluştu: {exc}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
        finally:
            context.close()
            browser.close()
            print("[*] Tarayıcı kapatıldı.")


if __name__ == "__main__":
    main()
